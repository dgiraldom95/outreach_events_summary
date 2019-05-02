import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
from Program import Program

CBOutreachFile = 'CB Outreach Education Events.xlsx'
outputFile = 'output.xlsx'
templateFile = 'template.xlsx'

firstRowInput = 4
lastRowInput = 4
lastYearData = 0
lastMonthData = 0

firstRowOutput = 8
lastRowOutput = 8
monthStartOutput = 6  # F
startYear = 2018
numberOfYears = 3
colsPerYear = 19

cInitiative = 'A'
cStrategy = 'B'
cActivity = 'C'
cProgram = 'D'
cDate = 'G'
cNumPeopleUnique = 'W'
cNumPeople = 'X'

data = {}  # data[initiative][strategy][activity][program]


def findLastRowInput(ws: Worksheet):
    i = firstRowInput
    global lastRowInput
    while ws['A' + str(i)].value is not None:
        i += 1
    lastRowInput = i - 1
    return lastRowInput


def findLastRowOutput(ws: Worksheet):
    i = firstRowOutput
    global lastRowOutput
    while ws['C' + str(i)].value is not None:
        i += 1
    lastRowOutput = i - 1
    return lastRowOutput


def returnProgramsInDictionary(dict):
    programList = []
    for key, value in dict.items():
        if isinstance(value, Program):
            programList.append(value)
        else:
            programList += returnProgramsInDictionary(value)
    return programList


def colorNext3Rows(ws: Worksheet, startRow):
    color = openpyxl.styles.Color(rgb='f0f0f0')
    for index, row in enumerate(ws.iter_rows(min_row=startRow, max_row=startRow + 2)):
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=color)


# Writes number of events, increase people served, number people served
def writeRowLabels(ws, row):
    ws['C' + str(row)] = 'Number of events'
    ws['C' + str(row + 1)] = 'Increase # of people served - Unduplicated (unique)'
    ws['C' + str(row + 1)].alignment = Alignment(wrapText=True)
    ws['C' + str(row + 2)] = 'Number of people served (Encounters)'
    ws['C' + str(row + 2)].alignment = Alignment(wrapText=True)


def writeAggregateFunctions(ws: Worksheet, startCol, firstRow, lastRow):
    if lastRow == None:
        lastRow = findLastRowOutput(ws)
    for i in range(firstRow, lastRow + 1):
        # Q1
        ws[get_column_letter(startCol) + str(i)] = '=SUM(%s%d:%s%d)' % (get_column_letter(startCol - 12),
                                                                        i,
                                                                        get_column_letter(startCol - 12 + 2),
                                                                        i)
        # Q2
        ws[get_column_letter(startCol + 1) + str(i)] = '=SUM(%s%d:%s%d)' % (get_column_letter(startCol - 12 + 3),
                                                                            i,
                                                                            get_column_letter(startCol - 12 + 5),
                                                                            i)
        # 6 month
        ws[get_column_letter(startCol + 2) + str(i)] = '=SUM(%s%d:%s%d)' % (get_column_letter(startCol - 12),
                                                                            i,
                                                                            get_column_letter(startCol - 12 + 5),
                                                                            i)
        # Q3
        ws[get_column_letter(startCol + 3) + str(i)] = '=SUM(%s%d:%s%d)' % (get_column_letter(startCol - 12 + 6),
                                                                            i,
                                                                            get_column_letter(startCol - 12 + 8),
                                                                            i)
        # Q4
        ws[get_column_letter(startCol + 4) + str(i)] = '=SUM(%s%d:%s%d)' % (get_column_letter(startCol - 12 + 9),
                                                                            i,
                                                                            get_column_letter(startCol - 12 + 11),
                                                                            i)
        # Year
        ws[get_column_letter(startCol + 5) + str(i)] = '=SUM(%s%d:%s%d)' % (get_column_letter(startCol - 12),
                                                                            i,
                                                                            get_column_letter(startCol - 12 + 11),
                                                                            i)


def applyBorders(ws: Worksheet):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for index, row in enumerate(ws.iter_rows()):
        for cell in row:
            cell.border = thin_border


def applyStrategyBorders(ws: Worksheet):
    top_thick_border = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thick'),
                              bottom=Side(style='thin'))
    bottom_thick_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thick'))

    for index, row in enumerate(ws.iter_rows()):
        if ws['A' + str(index + 1)].value is not None and (index + 1) >= firstRowOutput:
            for cell in row:
                cell.border = top_thick_border
            rowBottom = ws[index + 2 + 1]
            for cell in rowBottom:
                cell.border = bottom_thick_border


def getLastMonthAndYearWithEntries():
    global lastYearData
    global lastMonthData

    programs = returnProgramsInDictionary(data)

    lastMonth = 0
    lastYear = 0
    for prog in programs:
        month, year = prog.getLastYearAndMonthWithEntries()
        if (year >= lastYear and month > lastMonth) or (year > lastYear):
            lastMonth, lastYear = month, year

    lastYearData = lastYear
    lastMonthData = lastMonth


def load_data():
    wb = load_workbook(CBOutreachFile)
    ws = wb['Outreach Events']

    findLastRowInput(ws)

    for i in range(firstRowInput, lastRowInput + 1):
        global data
        row = str(i)
        initiative = ws[cInitiative + row].value
        strategy = ws[cStrategy + row].value
        activity = ws[cActivity + row].value
        program = ws[cProgram + row].value

        if program is None:
            continue
        else:
            program = program.strip()

        date = ws[cDate + row].value  # type: datetime

        numPeople = ws[cNumPeople + row].value
        numPeopleUnique = ws[cNumPeopleUnique + row].value

        if initiative not in data:
            data[initiative] = {}
        if strategy not in data[initiative]:
            data[initiative][strategy] = {}
        if activity not in data[initiative][strategy]:
            data[initiative][strategy][activity] = {}
        if program not in data[initiative][strategy][activity]:
            data[initiative][strategy][activity][program] = Program(initiative, strategy, activity, program)

        programInstance = data[initiative][strategy][activity][program]  # type: Program
        programInstance.addEvent(date, numPeople, numPeopleUnique)

    getLastMonthAndYearWithEntries()


def writeProgramCellsForYear(ws, row, program: Program, year):
    startCol = monthStartOutput + (year - startYear) * colsPerYear
    values = program.getMonthDict(year)

    lastMonth = 12
    if year == lastYearData and lastMonthData < lastMonth:
        lastMonth = lastMonthData

    for month in range(1, lastMonth + 1):
        ws[get_column_letter(startCol + month - 1) + str(row)] = values[month - 1]['numEvents']
        ws[get_column_letter(startCol + month - 1) + str(row + 1)] = values[month - 1]['numPeopleUnique']
        ws[get_column_letter(startCol + month - 1) + str(row + 2)] = values[month - 1]['numPeople']


# Writes the sum of the values for each month of all programs inside the given dictionary
def writeCellsForTotalsInDictionary(ws, row, year, dataDict):
    startCol = monthStartOutput + (year - startYear) * colsPerYear
    valuesToWrite = getTotalsInDictionaryForEachMonth(dataDict)

    lastMonth = 12
    if year == lastYearData and lastMonthData < lastMonth:
        lastMonth = lastMonthData

    for month in range(1, lastMonth + 1):
        events = valuesToWrite['%d-%d' % (year, month)]['numEvents']
        peopleUnique = valuesToWrite['%d-%d' % (year, month)]['numPeopleUnique']
        people = valuesToWrite['%d-%d' % (year, month)]['numPeople']

        ws[get_column_letter(startCol + month - 1) + str(row)] = events
        ws[get_column_letter(startCol + month - 1) + str(row + 1)] = peopleUnique
        ws[get_column_letter(startCol + month - 1) + str(row + 2)] = people


# Returns a dictionary (by year and month)with the total number of events, unique and people inside the given dictionary
def getTotalsInDictionaryForEachMonth(dictionary):
    totalsDict = {}
    programs = returnProgramsInDictionary(dictionary)
    for year in range(startYear, startYear + numberOfYears):
        for month in range(1, 13):
            numEvents = 0
            numPeopleUnique = 0
            numPeople = 0

            for program in programs:
                numEvents += program.events[str(year) + '-' + str(month)]['numEvents']
                numPeopleUnique += program.events[str(year) + '-' + str(month)]['numPeopleUnique']
                numPeople += program.events[str(year) + '-' + str(month)]['numPeople']

            totalsForMonth = {'numEvents': numEvents,
                              'numPeopleUnique': numPeopleUnique,
                              'numPeople': numPeople}
            totalsDict[str(year) + '-' + str(month)] = totalsForMonth

    return totalsDict


def writeData():
    for initiative in data:
        wb = load_workbook('template.xlsx')

        # Write Strategy
        strategyNum = 0
        for strategy in data[initiative]:
            strategyNum += 1
            template = wb['template']
            ws = wb.copy_worksheet(template)
            ws.title = '%s' % (strategy)
            ws['A3'] = 'Strategy: %s' % (strategy)

            row = firstRowOutput

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.freeze_panes = ws['D3']

            # Write info for each activity
            for activity in data[initiative][strategy]:
                # Write Activity Name
                ws['A' + str(row)] = activity
                ws['A' + str(row)].font = Font(bold=True)
                ws['A' + str(row)].alignment = Alignment(wrapText=True)
                ws.merge_cells('A%d:A%d' % (row, row + 2))

                # Write totals for activity
                ws['C' + str(row)] = 'Number of events'
                ws['C' + str(row + 1)] = 'Increase # of people served - Unduplicated (unique)'
                ws['C' + str(row + 1)].alignment = Alignment(wrapText=True)
                ws['C' + str(row + 2)] = 'Number of people served (Encounters)'
                ws['C' + str(row + 2)].alignment = Alignment(wrapText=True)

                for year in range(startYear, lastYearData + 1):
                    writeCellsForTotalsInDictionary(ws, row, year, data[initiative][strategy][activity])
                row += 3

                # Write header info for each program
                for program in data[initiative][strategy][activity]:
                    programInst = data[initiative][strategy][activity][program]  # type: Program

                    # Number of Events
                    ws['B' + str(row)] = program
                    ws['B' + str(row)].alignment = Alignment(wrapText=True)
                    ws['C' + str(row)] = 'Number of events'

                    ws['B' + str(row + 1)] = program
                    ws['B' + str(row + 1)].alignment = Alignment(wrapText=True)
                    ws['C' + str(row + 1)] = 'Increase # of people served - Unduplicated (unique)'
                    ws['C' + str(row + 1)].alignment = Alignment(wrapText=True)

                    ws['B' + str(row + 2)] = program
                    ws['B' + str(row + 2)].alignment = Alignment(wrapText=True)
                    ws['C' + str(row + 2)] = 'Number of people served (Encounters)'
                    ws['C' + str(row + 2)].alignment = Alignment(wrapText=True)

                    # Write info for event, iterate through years
                    for year in range(startYear, lastYearData + 1):
                        writeProgramCellsForYear(ws, row, programInst, year)

                    if row % 2 == 0:
                        colorNext3Rows(ws, row)

                    row += 3

            # Writes totals for the strategy
            writeRowLabels(ws, 3)
            for year in range(startYear, lastYearData + 1):
                writeCellsForTotalsInDictionary(ws, 3, year, data[initiative][strategy])
                writeAggregateFunctions(ws, monthStartOutput + 12 + colsPerYear * (year - startYear), 3, 5)

            for year in range(startYear, lastYearData + 1):
                writeAggregateFunctions(ws, monthStartOutput + 12 + colsPerYear * (year - startYear), firstRowOutput,
                                        None)

            applyBorders(ws)
            applyStrategyBorders(ws)
        wb.save('%s Summary.xlsx' % initiative)


if __name__ == '__main__':
    load_data()
    writeData()
